import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment, numbers
from openpyxl.utils import get_column_letter
import os
import io
import base64
import re
import tempfile
import shutil
import atexit # For cleanup
from datetime import datetime
from flask import Flask, render_template, request, redirect, url_for, flash
from werkzeug.utils import secure_filename
import logging

# Configure basic logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# --- Flask App Setup ---
app = Flask(__name__)
# IMPORTANT: Change this to a random secret key for production
app.config['SECRET_KEY'] = 'your-very-secret-and-random-key-change-me' # CHANGE THIS!
# Define a temporary directory for uploads within the instance folder
UPLOAD_FOLDER = os.path.join(app.instance_path, 'uploads')
# Ensure the instance folder and upload folder exist
try:
    os.makedirs(app.instance_path, exist_ok=True)
    os.makedirs(UPLOAD_FOLDER, exist_ok=True)
    logging.info(f"Upload folder created/ensured at: {UPLOAD_FOLDER}")
except OSError as e:
    logging.error(f"Could not create upload folder: {e}")
    # Depending on severity, you might want to exit or handle this differently
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024 # 16 MB limit for uploads

# ========== PASTE YOUR BASE64 MASTER FILE HERE ==========
# Replace the placeholder comment and the empty string below
# with your actual `TEMPLATE_DATA = """..."""` assignment.
# DO NOT leave it empty or the app will fail.
TEMPLATE_DATA = "" # PASTE YOUR ACTUAL BASE64 STRING HERE!

# --- Global variable for the decoded template path ---
_temp_template_path = None

def decode_master_template():
    """Decodes the base64 template string into a temporary Excel file path."""
    global _temp_template_path # To store the path globally
    if _temp_template_path and os.path.exists(_temp_template_path):
        logging.info("Temporary template already exists.")
        return _temp_template_path # Return existing path if already decoded

    try:
        if not TEMPLATE_DATA:
             # Added specific check for empty string after user prompt
            raise ValueError("TEMPLATE_DATA is empty. Please paste your base64 string into app.py.")
        if "[BASE64_TEMPLATE_DATA_HERE]" in TEMPLATE_DATA:
             # This check might be redundant if the above check works, but keep for safety
            raise ValueError("No valid base64 template data provided. Please replace the placeholder in app.py.")

        logging.info("Decoding base64 template data...")
        decoded_bytes = base64.b64decode(TEMPLATE_DATA)

        # Create the temp file within the instance folder for better permission handling on Render
        temp_dir = os.path.join(app.instance_path, 'temp_templates')
        os.makedirs(temp_dir, exist_ok=True)

        # Use NamedTemporaryFile correctly within the temp_dir
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx", prefix="template_", dir=temp_dir) as temp_excel_file:
            template_path = temp_excel_file.name
            temp_excel_file.write(decoded_bytes)
        _temp_template_path = template_path # Store the path globally
        logging.info(f"Decoded template saved to temporary file: {template_path}")
        return template_path
    except base64.binascii.Error as b64_error:
        logging.error(f"Error decoding base64 data: {b64_error}. Ensure the string is correct.")
        raise Exception(f"Error decoding base64 data: {b64_error}. Ensure the string is correct.")
    except ValueError as ve:
        logging.error(f"Value error during template decoding: {ve}")
        raise ve
    except Exception as e:
        logging.error(f"Error decoding/saving temporary template file: {e}")
        # Clean up potentially created file if error occurs after creation
        # Check if template_path was assigned before error
        if 'template_path' in locals() and template_path and os.path.exists(template_path):
            try:
                os.remove(template_path)
            except OSError:
                pass
        # Also check global variable in case error happened later
        elif '_temp_template_path' in globals() and _temp_template_path and os.path.exists(_temp_template_path):
             try: os.remove(_temp_template_path)
             except OSError: pass

        _temp_template_path = None # Reset global path on error
        raise Exception(f"Error decoding/saving temporary template file: {e}")


def cleanup_temp_template_on_exit():
    """Function to clean up the temporary template file on application exit."""
    global _temp_template_path
    if _temp_template_path and os.path.exists(_temp_template_path):
        try:
            os.remove(_temp_template_path)
            logging.info(f"Cleaned up temporary template file: {_temp_template_path}")
        except OSError as e:
            logging.warning(f"Could not remove temp template file {_temp_template_path} on exit: {e}")
    # Clean up the temp directory as well, if empty
    temp_dir = os.path.join(app.instance_path, 'temp_templates')
    if os.path.exists(temp_dir):
        try:
            if not os.listdir(temp_dir): # Check if directory is empty
                 os.rmdir(temp_dir)
                 logging.info(f"Cleaned up empty temp template directory: {temp_dir}")
        except OSError as e:
             logging.warning(f"Could not remove temp template directory {temp_dir} on exit: {e}")


# --- Register cleanup function to run when the application exits ---
atexit.register(cleanup_temp_template_on_exit)

# --- Financial Processor Class (Adapted for Web) ---
class FinancialStatementProcessor:
    # Keep most methods as they are, just add data extraction
    def __init__(self, template_path):
        self.template_path = template_path
        self.wb_template_structure = None # Store the initial template structure
        try:
            # Load the template structure once during initialization
            self.wb_template_structure = self.load_template_from_path(data_only=False)
            self.validate_template_sheets(self.wb_template_structure)
        except Exception as e:
            logging.error(f"Processor Initialization failed: {e}")
            # No need to call cleanup_temp_template here, atexit handles it
            raise e # Re-raise to prevent app start if processor fails

    def load_template_from_path(self, data_only=False):
        try:
            if not os.path.exists(self.template_path):
                raise FileNotFoundError(f"Template not found: {self.template_path}")
            # Load a fresh copy each time processing happens if modifying it
            wb = load_workbook(self.template_path, data_only=data_only)
            logging.info(f"Loaded template structure from: {self.template_path} (data_only={data_only})")
            return wb
        except Exception as e:
            logging.error(f"Error loading template: {e}")
            raise Exception(f"Error loading template: {e}")

    def validate_template_sheets(self, wb_to_check):
        required = ["Income Statement", "Balance Sheet", "Cash Flow Statement"]
        available_sheets = wb_to_check.sheetnames
        for sheet in required:
            if sheet not in available_sheets:
                logging.error(f"Template sheet validation failed. Missing: '{sheet}'. Available: {available_sheets}")
                raise ValueError(f"Required sheet '{sheet}' missing in template. Available: {available_sheets}")
        logging.info("Template sheets validated successfully.")

    def load_csv(self, file_path, sheet_name):
        try:
            logging.info(f"Loading CSV: {os.path.basename(file_path)} for sheet {sheet_name}")
            df = pd.read_csv(file_path)
            logging.info(f"Successfully loaded CSV: {os.path.basename(file_path)}")
            return df
        except FileNotFoundError:
            logging.error(f"CSV not found: {file_path}")
            raise FileNotFoundError(f"CSV file not found: {os.path.basename(file_path)}")
        except pd.errors.EmptyDataError:
            logging.warning(f"CSV file is empty: {file_path}")
            # Return an empty DataFrame instead of raising an error immediately
            return pd.DataFrame()
        except Exception as e:
            logging.error(f"Error reading CSV {file_path}: {e}")
            raise Exception(f"Error reading CSV {os.path.basename(file_path)}: {e}")

    def clean_data(self, df, sheet_name):
        if df.empty:
             logging.warning(f"Skipping cleaning for empty DataFrame: {sheet_name}")
             return df
        logging.info(f"Cleaning data for sheet: {sheet_name}")
        df.columns = df.columns.str.strip()
        if df.columns.tolist() and len(df.columns) >= 1:
            try:
                # Ensure the first column exists before trying to access iloc[:, 0]
                if df.shape[1] > 0:
                    df.iloc[:, 0] = df.iloc[:, 0].astype(str).str.replace(r"^\s+|\s+$|\t", "", regex=True)
                else:
                     logging.warning(f"DataFrame for {sheet_name} has no columns, skipping first column cleaning.")
            except Exception as e:
                logging.warning(f"Warning: Cleaning the first column failed for {sheet_name}: {e}")
        # Drop rows where ALL columns are NaN
        df.dropna(how='all', inplace=True)
        logging.info(f"Finished cleaning data for sheet: {sheet_name}. Shape: {df.shape}")
        return df

    def append_data_to_excel(self, df, wb, sheet_name, start_row):
        if sheet_name not in wb.sheetnames:
             logging.error(f"Sheet '{sheet_name}' not found in workbook during append.")
             raise ValueError(f"Sheet '{sheet_name}' not found.")

        ws = wb[sheet_name]
        target_start_row = start_row
        # Find the first truly empty row starting from 'start_row'
        current_row_check = start_row
        while current_row_check <= ws.max_row + 5: # Check a bit beyond max_row just in case
            row_has_data = False
            # Check only first few columns for performance maybe? Let's check all for now.
            for cell in ws[current_row_check]:
                if cell.value is not None and str(cell.value).strip() != "":
                    row_has_data = True
                    break
            if not row_has_data:
                 # Check if the next few rows are also empty to confirm end of data
                 all_next_empty = True
                 for next_row_offset in range(1, 4):
                    next_row_idx = current_row_check + next_row_offset
                    # Only check within actual sheet bounds if possible
                    if next_row_idx <= ws.max_row:
                         for cell in ws[next_row_idx]:
                              if cell.value is not None and str(cell.value).strip() != "":
                                   all_next_empty = False
                                   break
                    if not all_next_empty: break # Stop checking next rows if one has data
                 if all_next_empty:
                     target_start_row = current_row_check
                     break # Found the start of the empty block

            current_row_check += 1

        # If loop finished without finding an empty block, append after last row with data
        if current_row_check > ws.max_row + 5:
             actual_max_row = 1 # Default to 1 if sheet is completely empty initially
             for row in ws.iter_rows(min_row=1, max_col=ws.max_column, max_row=ws.max_row):
                  row_contains_data = False
                  for cell in row:
                       if cell.value is not None and str(cell.value).strip() != "":
                            row_contains_data = True
                            break
                  if row_contains_data:
                       actual_max_row = row[0].row # Get row index from the first cell of the row
             target_start_row = actual_max_row + 1
             logging.info(f"Could not find empty block, determined last data row as {actual_max_row}, appending from {target_start_row}")


        logging.info(f"Determined append start row for '{sheet_name}' as {target_start_row}")

        if df.empty:
            logging.warning(f"DataFrame for '{sheet_name}' is empty. Skipping append.")
            return # Don't try to append an empty dataframe

        logging.info(f"Appending {len(df)} rows to '{sheet_name}' starting at row {target_start_row}")

        for r_offset, row_data_tuple in enumerate(df.itertuples(index=False, name=None)):
            current_ws_row = target_start_row + r_offset
            # Ensure we don't write beyond reasonable column limits
            max_cols_to_write = min(len(row_data_tuple), 50) # Limit writing width
            for c_offset, value in enumerate(row_data_tuple[:max_cols_to_write], start=1):
                # Defensive check for row/column indices (should be okay here but safe)
                if current_ws_row > openpyxl.worksheet.worksheet.Worksheet.max_row or c_offset > openpyxl.worksheet.worksheet.Worksheet.max_column:
                    logging.warning(f"Skipping write beyond max excel limits at R{current_ws_row}C{c_offset} in sheet {sheet_name}")
                    continue

                cell_to_write = ws.cell(row=current_ws_row, column=c_offset)
                try:
                    if pd.isna(value):
                        cell_to_write.value = None
                    elif isinstance(value, str):
                        # Attempt to convert numeric strings
                        cleaned_value = value.replace(',', '').strip()
                        if not cleaned_value: # Handle empty strings
                             cell_to_write.value = None
                             continue
                        try:
                            # Check for parentheses indicating negative numbers
                            is_negative = False
                            if cleaned_value.startswith('(') and cleaned_value.endswith(')'):
                                cleaned_value = cleaned_value[1:-1]
                                is_negative = True

                            if '.' in cleaned_value:
                                num_value = float(cleaned_value)
                            else:
                                num_value = int(cleaned_value)

                            cell_to_write.value = -num_value if is_negative else num_value
                            cell_to_write.number_format = numbers.FORMAT_NUMBER_00 # Apply basic format
                        except ValueError:
                            # If conversion fails, keep as string
                            cell_to_write.value = value
                    elif isinstance(value, (int, float)):
                         cell_to_write.value = value
                         cell_to_write.number_format = numbers.FORMAT_NUMBER_00 # Apply basic format
                    elif isinstance(value, datetime):
                         cell_to_write.value = value
                         cell_to_write.number_format = numbers.FORMAT_DATE_YYYYMMDD2 # Apply date format
                    else:
                         # Try converting other types to string as fallback
                         cell_to_write.value = str(value)

                except Exception as cell_write_error:
                    logging.error(f"Error writing cell (R{current_ws_row}, C{c_offset}) for sheet '{sheet_name}': {cell_write_error}. Value: {repr(value)}")
                    try:
                        cell_to_write.value = str(value) # Fallback to string representation
                    except:
                        cell_to_write.value = "WRITE_ERROR" # Final fallback

        # Apply alignment formatting after appending all data for this sheet
        self.apply_formatting(ws, target_start_row, len(df))


    def apply_formatting(self, ws, start_row, num_rows):
        if num_rows <= 0: return
        end_row = start_row + num_rows - 1
        logging.info(f"Applying alignment formatting to '{ws.title}' rows {start_row}-{end_row}")
        max_col_to_format = min(ws.max_column + 1, 50) # Limit formatting width

        for row_idx in range(start_row, end_row + 1):
            # Check row existence defensively
            if row_idx > openpyxl.worksheet.worksheet.Worksheet.max_row:
                 logging.warning(f"Attempted to format non-existent row {row_idx} in sheet '{ws.title}'.")
                 continue
            for col_idx in range(1, max_col_to_format):
                 # Check column existence defensively
                if col_idx > openpyxl.worksheet.worksheet.Worksheet.max_column:
                     continue
                try:
                    cell = ws.cell(row=row_idx, column=col_idx)
                    # Apply alignment
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                    # Number formats are now applied during append, so this only does alignment
                except Exception as e:
                    cell_coord = f"R{row_idx}C{col_idx}"
                    try:
                        cell_coord = ws.cell(row=row_idx, column=col_idx).coordinate
                    except: pass # Ignore if coordinate fails
                    logging.warning(f"Alignment formatting error in sheet '{ws.title}' at cell {cell_coord}. Error: {e}")

    def update_formulas(self, wb, data_length_map, formula_config):
        logging.info("Starting formula update process...")
        max_end_row_map = {}

        # Determine the actual end row for data in each relevant sheet
        for sheet_name, details in formula_config.items():
            data_start_row_config = details['adjust_rows_from'] # The row where data STARTS
            if sheet_name in data_length_map and data_length_map[sheet_name] > 0:
                 # Find the actual start row of the appended data
                ws_check = wb[sheet_name]
                first_data_row = -1
                current_row_check = data_start_row_config
                # Check rows where data could have been appended
                # Increase check range slightly beyond expected end based on data length
                max_check_row = data_start_row_config + data_length_map[sheet_name] + 5
                max_check_row = min(max_check_row, ws_check.max_row + 5) # Don't check excessively far

                while current_row_check <= max_check_row:
                    # Check if *any* cell in the row has data
                    row_has_data = any(cell.value is not None and str(cell.value).strip() != "" for cell in ws_check[current_row_check])
                    if row_has_data and first_data_row == -1:
                         # Potential start row found
                         # Check if the row *before* the configured start was actually empty, suggesting config row is correct start
                         if current_row_check == data_start_row_config:
                              prev_row_empty = True
                              if current_row_check > 1:
                                   prev_row_empty = all(cell.value is None or str(cell.value).strip() == "" for cell in ws_check[current_row_check - 1])
                              if prev_row_empty:
                                   first_data_row = current_row_check
                                   # Don't break yet, keep checking in case of empty rows within data
                         else:
                              # Found data after configured start, likely the real start
                               first_data_row = current_row_check
                               # Don't break yet

                    elif not row_has_data and first_data_row != -1:
                         # Found an empty row *after* finding some data. Assume data ended just before this.
                         max_end_row_map[sheet_name] = current_row_check - 1
                         logging.info(f"Determined data range for '{sheet_name}': Rows {first_data_row} to {max_end_row_map[sheet_name]}")
                         break # Exit inner loop once end is found

                    current_row_check += 1
                else: # Loop finished without finding an empty row after data started
                    if first_data_row != -1:
                        # Data likely goes to the end of checked range or sheet max row
                        max_end_row_map[sheet_name] = first_data_row + data_length_map[sheet_name] - 1 # Calculate based on length
                        # Cap at actual sheet max row
                        max_end_row_map[sheet_name] = min(max_end_row_map[sheet_name], ws_check.max_row)
                        logging.info(f"Data seems contiguous for '{sheet_name}'. Determined range: Rows {first_data_row} to {max_end_row_map[sheet_name]}")
                    else:
                        # Could not find any data start row, use config start row for end calculation
                        max_end_row_map[sheet_name] = data_start_row_config + data_length_map[sheet_name] - 1
                        logging.warning(f"Could not find data start row for '{sheet_name}'. Using config start {data_start_row_config}. Calculated end row: {max_end_row_map[sheet_name]}")

            else:
                # If no data was appended, the "last row" for formula adjustment is effectively the row *before* data would start
                max_end_row_map[sheet_name] = data_start_row_config - 1
                logging.info(f"No data appended to '{sheet_name}'. Effective last row for formula adjustment: {max_end_row_map[sheet_name]}")

        # Now, adjust formulas based on the calculated max_end_row_map
        for sheet_name_formula_adjustments, details in formula_config.items():
            if sheet_name_formula_adjustments not in wb.sheetnames:
                logging.warning(f"Sheet '{sheet_name_formula_adjustments}' specified in formula_config not found in workbook. Skipping.")
                continue

            ws = wb[sheet_name_formula_adjustments]
            formula_range_str = details['range']
            logging.info(f"Processing formulas in '{sheet_name_formula_adjustments}' range {formula_range_str}")

            try:
                min_col_idx, min_row_idx, max_col_idx, max_row_idx = openpyxl.utils.range_boundaries(formula_range_str)
            except Exception as range_parse_error:
                logging.error(f"Error parsing formula range '{formula_range_str}' for sheet '{sheet_name_formula_adjustments}': {range_parse_error}. Skipping sheet.")
                continue

            # Regex specifically for VLOOKUP range adjustment (improved)
            # Handles 'Sheet Name'!A1:B10, SheetName!A1:B10, A1:B10, $A$1:$B$10 etc.
            vlookup_pattern = re.compile(
                 r"(VLOOKUP\s*\([^,]+,\s*)"  # Start of VLOOKUP, lookup value, comma (Group 1)
                 r"((?:'[^']+'!|[^'!]+!)?"  # Optional Sheet Prefix (quoted or unquoted) (Group 3 within Group 2)
                 r"(\$?[A-Za-z]+\$?\d+:\$?[A-Za-z]+\$?\d+))"  # The range A1:B10, $A$1:$B$10 (Group 4 within Group 2, total range is Group 2)
                 r"(\s*,)"  # Comma after range (Group 5)
            )

            def replace_vlookup_range(match):
                try:
                    vlookup_prefix = match.group(1)       # e.g., "VLOOKUP(A2,"
                    original_range_str = match.group(2)   # e.g., "'Sheet1'!$A$5:B10" or Data!A5:B10 or $C$5:$D$10
                    sheet_prefix = match.group(3) or ''   # e.g., "'Sheet1'!" or "Data!" or ""
                    range_only = match.group(4)           # e.g., "$A$5:B10" or "A5:B10"
                    vlookup_suffix_comma = match.group(5) # e.g., ","

                    # Parse the range_only part
                    range_match = re.match(r"(\$?[A-Za-z]+\$?)(\d+):(\$?[A-Za-z]+\$?)(\d+)", range_only)
                    if not range_match:
                         logging.warning(f"    Could not parse range part '{range_only}' from VLOOKUP. Skipping.")
                         return match.group(0) # Return original match
                    start_col_ref, start_row, end_col_ref, end_row_old = range_match.groups()

                    # Determine the target sheet for max_end_row lookup
                    target_sheet_name = sheet_name_formula_adjustments # Default to current sheet
                    if sheet_prefix:
                        # Extract sheet name, handling quotes
                        sheet_match = re.match(r"'([^']+)'!", sheet_prefix)
                        if sheet_match:
                            target_sheet_name = sheet_match.group(1)
                        else: # Unquoted sheet name
                             target_sheet_name = sheet_prefix[:-1] # Remove trailing '!'

                    if target_sheet_name in max_end_row_map:
                        new_end_row_num = max_end_row_map[target_sheet_name]
                        # Only adjust if the new end row makes sense (is at least the start row)
                        if new_end_row_num >= int(start_row):
                            # Reconstruct the range using original column refs and new end row
                            new_range = f"{start_col_ref}{start_row}:{end_col_ref}{new_end_row_num}"
                            # Reconstruct the full VLOOKUP part
                            reconstructed = f"{vlookup_prefix}{sheet_prefix}{new_range}{vlookup_suffix_comma}"
                            # logging.debug(f"    Replacing VLOOKUP range {original_range_str} -> {sheet_prefix}{new_range}")
                            return reconstructed
                        else:
                             logging.warning(f"    VLOOKUP adjustment skipped for {original_range_str} in {target_sheet_name}: new end row {new_end_row_num} is before start row {start_row}.")
                             return match.group(0) # Return original match if adjustment doesn't make sense
                    else:
                         logging.warning(f"    VLOOKUP adjustment skipped: Sheet '{target_sheet_name}' (from range {original_range_str}) not found in calculated max_end_row_map.")
                         return match.group(0) # Return original match if sheet not found

                except IndexError:
                    logging.error(f"    Regex group index error processing VLOOKUP match: {match.group(0)}. Skipping adjustment.")
                    return match.group(0)
                except Exception as e:
                    logging.error(f"    Unexpected error in replace_vlookup_range for {match.group(0)}: {e}. Skipping adjustment.")
                    return match.group(0)

            # Iterate through cells in the specified formula range
            for row_idx in range(min_row_idx, max_row_idx + 1):
                for col_idx in range(min_col_idx, max_col_idx + 1):
                    # Add boundary checks for safety
                    if row_idx > ws.max_row or col_idx > ws.max_column: continue
                    try:
                        cell = ws.cell(row=row_idx, column=col_idx)
                        if cell.data_type == 'f' and isinstance(cell.value, str) and cell.value.startswith('='):
                            original_formula = cell.value
                            # Apply VLOOKUP range adjustment
                            new_formula = vlookup_pattern.sub(replace_vlookup_range, original_formula)

                            # Add more formula adjustment patterns here if needed (e.g., SUM)

                            if new_formula != original_formula:
                                logging.info(f"Updating formula: {ws.title}!{cell.coordinate} From: '{original_formula}' To: '{new_formula}'")
                                cell.value = new_formula
                    except Exception as cell_proc_error:
                         cell_coord = f"R{row_idx}C{col_idx}"
                         try: cell_coord = ws.cell(row=row_idx, column=col_idx).coordinate
                         except: pass
                         logging.error(f"Error processing cell {cell_coord} in '{ws.title}' during formula update: {cell_proc_error}")

        logging.info("Formula update finished.")


    def _extract_data_from_workbook(self, wb, sheet_configs):
        """ Extracts data from specified sheets and ranges in the workbook. """
        extracted_data = {}
        logging.info("Starting data extraction from processed workbook.")
        for sheet_name, config in sheet_configs.items():
            if sheet_name not in wb.sheetnames:
                logging.warning(f"Sheet '{sheet_name}' not found in workbook for extraction.")
                extracted_data[sheet_name] = {'headers': [], 'data': []}
                continue

            ws = wb[sheet_name]
            data_range_str = config.get('display_range', None)
            header_row_num = config.get('header_row', 1) # Default to row 1 if not specified

            sheet_data = []
            headers = []

            if data_range_str:
                try:
                    min_col_idx, min_row_idx, max_col_idx, max_row_idx = openpyxl.utils.range_boundaries(data_range_str)

                    # Ensure max row doesn't exceed actual sheet dimensions
                    max_row_idx = min(max_row_idx, ws.max_row)
                    max_col_idx = min(max_col_idx, ws.max_column)


                    # --- Extract Headers ---
                    # Check if header row is valid and within sheet bounds
                    if header_row_num >= 1 and header_row_num <= max_row_idx :
                        actual_header_row = ws[header_row_num]
                        # Extract headers only within the specified column range
                        headers = [cell.value for idx, cell in enumerate(actual_header_row) if min_col_idx <= idx + 1 <= max_col_idx]
                        # Adjust data start row if headers were within the display range
                        if header_row_num >= min_row_idx:
                             min_row_idx = header_row_num + 1
                    else:
                         logging.warning(f"Header row {header_row_num} is outside the sheet bounds or display range for sheet '{sheet_name}'. No headers extracted.")
                         headers = [""] * (max_col_idx - min_col_idx + 1) # Placeholder headers


                    # --- Extract Data Rows ---
                    logging.info(f"Extracting data from '{sheet_name}' calculated range {get_column_letter(min_col_idx)}{min_row_idx}:{get_column_letter(max_col_idx)}{max_row_idx}")
                    # Ensure min_row_idx is not greater than max_row_idx after header adjustment
                    if min_row_idx <= max_row_idx:
                        for row_idx in range(min_row_idx, max_row_idx + 1):
                             # Extract row data only within the specified column range
                             row_data = [ws.cell(row=row_idx, column=col_idx).value for col_idx in range(min_col_idx, max_col_idx + 1)]

                             # Apply basic formatting for display (optional, can be done in Jinja2 too)
                             formatted_row = []
                             for cell_value in row_data:
                                 if isinstance(cell_value, (int, float)):
                                     # Basic number formatting for display
                                     try:
                                          # Simple comma format, 2 decimal places
                                          formatted_row.append(f"{cell_value:,.2f}")
                                     except (ValueError, TypeError):
                                          formatted_row.append(cell_value) # Fallback
                                 elif isinstance(cell_value, datetime):
                                     formatted_row.append(cell_value.strftime('%Y-%m-%d'))
                                 else:
                                     formatted_row.append(cell_value) # Keep strings, None, etc. as is

                             sheet_data.append(formatted_row)
                    else:
                         logging.info(f"No data rows to extract for sheet '{sheet_name}' after header processing (min_row > max_row).")


                except Exception as extract_error:
                    logging.error(f"Error extracting data from range '{data_range_str}' in sheet '{sheet_name}': {extract_error}", exc_info=True)
                    # Provide empty data on error for this sheet
                    headers = []
                    sheet_data = []
            else:
                 logging.warning(f"No 'display_range' specified for sheet '{sheet_name}'. Skipping data extraction.")


            extracted_data[sheet_name] = {'headers': headers, 'data': sheet_data}
            logging.info(f"Extracted {len(sheet_data)} rows of data with {len(headers)} headers for sheet '{sheet_name}'.")

        logging.info("Finished data extraction from workbook.")
        return extracted_data

    # Main processing method called by Flask
    def process_files_for_web(self, file_paths):
        """
        Processes uploaded CSV files using the template and returns extracted data
        suitable for web display. Doesn't save the final Excel file.
        """
        wb = None # Ensure wb is defined in this scope
        wb_data_only = None # Define wb_data_only
        temp_output_path = None # Keep track of temp file if created
        ticker_symbol = None # Initialize ticker_symbol

        try:
            if len(file_paths) != 3:
                raise ValueError("Please provide exactly 3 CSV file paths.")

            file_map = {}

            # --- File Classification ---
            logging.info("Classifying input files...")
            for file_path in file_paths:
                filename = os.path.basename(file_path)
                match = re.match(r"([A-Za-z0-9]+)_annual_(cash-flow|balance-sheet|financials)\.csv", filename, re.IGNORECASE)
                if not match:
                    raise ValueError(f"Invalid filename format: {filename}. Expected TICKER_annual_type.csv")

                current_ticker, sheet_type_raw = match.groups()
                current_ticker_upper = current_ticker.upper()

                if ticker_symbol is None:
                    ticker_symbol = current_ticker_upper
                elif ticker_symbol != current_ticker_upper:
                    raise ValueError(f"Ticker symbol mismatch in filenames: Expected '{ticker_symbol}', found '{current_ticker_upper}' in {filename}")

                stype = sheet_type_raw.lower()
                if stype == "financials" and 'income' not in file_map:
                    file_map['income'] = file_path
                elif stype == "balance-sheet" and 'balance' not in file_map:
                    file_map['balance'] = file_path
                elif stype == "cash-flow" and 'cashflow' not in file_map:
                    file_map['cashflow'] = file_path
                else:
                     # Handle duplicate types
                     raise ValueError(f"Duplicate file type '{stype}' found or invalid type for filename {filename}")

            if len(file_map) != 3:
                 missing = {'income', 'balance', 'cashflow'} - file_map.keys()
                 # Map internal keys back to expected file types for user message
                 type_map = {'income': 'financials', 'balance': 'balance-sheet', 'cashflow': 'cash-flow'}
                 missing_types = [type_map[m] for m in missing]
                 raise ValueError(f"Missing required file types: {', '.join(missing_types)}")
            logging.info(f"Files classified successfully for ticker: {ticker_symbol}")

            # --- Load and Clean Data ---
            logging.info("Loading and cleaning CSV data...")
            income_df = self.clean_data(self.load_csv(file_map['income'], "Income Statement"), "Income Statement")
            balance_df = self.clean_data(self.load_csv(file_map['balance'], "Balance Sheet"), "Balance Sheet")
            cash_flow_df = self.clean_data(self.load_csv(file_map['cashflow'], "Cash Flow Statement"), "Cash Flow Statement")

            # --- Prepare In-Memory Workbook ---
            logging.info("Creating temporary workbook from template...")
            # Create a unique temp file name
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx", prefix=f"{ticker_symbol}_proc_", dir=app.config['UPLOAD_FOLDER']) as temp_wb_file:
                 temp_output_path = temp_wb_file.name
            shutil.copy(self.template_path, temp_output_path)
            logging.info(f"Copied template to temporary path: {temp_output_path}")

            # Load the temporary workbook for modification (formulas need data_only=False)
            wb = load_workbook(temp_output_path, data_only=False)
            logging.info("Loaded temporary workbook for processing.")

            # --- Append Data ---
            logging.info("Appending data to temporary workbook...")
            sheet_append_info = {
                "Income Statement": 10,
                "Balance Sheet": 7,
                "Cash Flow Statement": 9
            }
            self.append_data_to_excel(income_df, wb, "Income Statement", sheet_append_info["Income Statement"])
            self.append_data_to_excel(balance_df, wb, "Balance Sheet", sheet_append_info["Balance Sheet"])
            self.append_data_to_excel(cash_flow_df, wb, "Cash Flow Statement", sheet_append_info["Cash Flow Statement"])

            # --- Update Formulas ---
            logging.info("Updating formulas in temporary workbook...")
            data_lengths = {
                "Income Statement": len(income_df),
                "Balance Sheet": len(balance_df),
                "Cash Flow Statement": len(cash_flow_df)
            }
            formula_config = {
                "Income Statement":    {'range': 'C2:L8', 'adjust_rows_from': sheet_append_info["Income Statement"]},
                "Balance Sheet":       {'range': 'B2:K5', 'adjust_rows_from': sheet_append_info["Balance Sheet"]},
                "Cash Flow Statement": {'range': 'C2:L5', 'adjust_rows_from': sheet_append_info["Cash Flow Statement"]}
            }
            self.update_formulas(wb, data_lengths, formula_config)

            # --- Specific Formatting ---
            logging.info("Applying specific formatting to Cash Flow Statement rows 2 & 3...")
            try:
                if "Cash Flow Statement" in wb.sheetnames:
                    cf_ws = wb["Cash Flow Statement"]
                    three_decimal_format = "0.000"
                    # Determine max column dynamically but cap it reasonably
                    max_col_to_format = min(cf_ws.max_column + 1, 27) # Cap at Z
                    for row_idx in [2, 3]:
                        if row_idx <= cf_ws.max_row:
                            for col_idx in range(3, max_col_to_format): # Start from column C (3)
                               if col_idx <= cf_ws.max_column:
                                    cell = cf_ws.cell(row=row_idx, column=col_idx)
                                    # Check if cell contains a number before formatting
                                    if isinstance(cell.value, (int, float)):
                                         cell.number_format = three_decimal_format
                                    # else: Don't format non-numeric cells
                else:
                    logging.warning("Cash Flow Statement sheet not found for specific formatting.")
            except Exception as fmt_error:
                logging.warning(f"Could not apply specific formatting to Cash Flow rows 2-3: {fmt_error}")

            # --- Save temporary workbook to calculate formulas ---
            logging.info("Saving temporary workbook to calculate formulas...")
            wb.save(temp_output_path)
            wb.close() # Close the workbook object
            wb = None # Reset wb variable
            logging.info("Temporary workbook saved and closed.")

            # --- Reload Workbook with Calculated Values ---
            logging.info("Reloading temporary workbook with calculated values (data_only=True)...")
            wb_data_only = load_workbook(temp_output_path, data_only=True)
            logging.info("Reloaded workbook with data_only=True.")


            # --- Extract Data for Display ---
            logging.info("Extracting data for web display...")
            # Define which parts of each sheet to display - ADJUST THESE RANGES AS NEEDED
            sheet_display_configs = {
                 # Ranges should cover headers and potential data area
                "Income Statement":  {'display_range': 'A1:L40', 'header_row': 9}, # Adjusted range potentially
                "Balance Sheet":     {'display_range': 'A1:K50', 'header_row': 6}, # Adjusted range potentially
                "Cash Flow Statement":{'display_range': 'A1:L50', 'header_row': 8} # Adjusted range potentially
            }
            processed_data = self._extract_data_from_workbook(wb_data_only, sheet_display_configs)

            wb_data_only.close() # Close the data_only workbook
            wb_data_only = None # Reset
            logging.info("Closed data_only workbook.")

            logging.info("Processing for web display complete.")
            return {'ticker': ticker_symbol, 'sheets': processed_data}

        except Exception as e:
            logging.error(f"Error during web processing: {e}", exc_info=True) # Log traceback
            # Ensure workbooks are closed if they were opened
            if wb is not None:
                try: wb.close()
                except: pass
            if wb_data_only is not None:
                try: wb_data_only.close()
                except: pass
            raise # Re-raise the exception for Flask handler

        finally:
             # --- Cleanup Temporary Processing Workbook File ---
             if temp_output_path and os.path.exists(temp_output_path):
                 try:
                      os.remove(temp_output_path)
                      logging.info(f"Cleaned up temporary processing workbook: {temp_output_path}")
                 except OSError as e:
                      logging.warning(f"Could not remove temporary processing workbook {temp_output_path}: {e}")


# --- Global Processor Instance ---
# Initialize processor when the app starts
try:
    # Decode the template first
    _template_path_on_startup = decode_master_template()
    if _template_path_on_startup:
        # Create the processor instance
        processor = FinancialStatementProcessor(_template_path_on_startup)
        logging.info("FinancialStatementProcessor initialized successfully.")
    else:
         # Should not happen if decode_master_template raises Exception on failure
         logging.fatal("FATAL: decode_master_template returned None without raising error.")
         processor = None
except Exception as e:
    logging.fatal(f"FATAL: Could not initialize FinancialStatementProcessor: {e}", exc_info=True)
    # If the processor fails to initialize, the app can't run.
    processor = None # Ensure processor is None if init fails


# --- Flask Routes ---
@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        # Check if processor initialized correctly
        if processor is None:
             flash('Application Error: The statement processor could not be initialized. Please check logs.', 'danger')
             logging.error("Processing attempt failed because processor is None.")
             # Maybe redirect to an error page or just back to index
             return redirect(url_for('index'))

        # --- File Upload Handling ---
        if 'csv_files' not in request.files:
            flash('No file part in the request.', 'warning')
            return redirect(request.url)

        files = request.files.getlist('csv_files')

        # Check if any file was actually selected
        if not files or all(f.filename == '' for f in files):
             flash('No files selected. Please select the three required CSV files.', 'warning')
             return redirect(request.url)

        if len(files) != 3:
            flash(f'Please select exactly 3 CSV files. You selected {len(files)}.', 'warning')
            return redirect(request.url)

        saved_files = []
        temp_upload_dir = app.config['UPLOAD_FOLDER']
        try:
            # Save files temporarily
            for file in files:
                 # Double check file object and filename
                 if file and file.filename:
                    if file.filename.lower().endswith('.csv'):
                        # Sanitize filename
                        filename = secure_filename(file.filename)
                        if not filename: # Handle cases where secure_filename returns empty string
                            filename = f"upload_{datetime.now().timestamp()}.csv" # Fallback name
                        filepath = os.path.join(temp_upload_dir, filename)
                        file.save(filepath)
                        saved_files.append(filepath)
                        logging.info(f"Saved uploaded file: {filepath}")
                    else:
                        flash(f'Invalid file type: "{file.filename}". Only CSV files are allowed.', 'danger')
                        raise ValueError("Invalid file type uploaded.") # Raise error to trigger cleanup
                 else:
                    # Handle case where one of the file inputs might be empty/invalid
                     flash('One of the file inputs was empty or invalid.', 'danger')
                     raise ValueError("Empty or invalid file input.") # Raise error to trigger cleanup

            # --- Process Files ---
            logging.info("Calling processor.process_files_for_web...")
            results_data = processor.process_files_for_web(saved_files)
            logging.info("Processing successful.")

            # --- Render Results ---
            # Don't flash success here, the results page is the success indicator
            return render_template('results.html', results=results_data)

        except ValueError as ve:
             flash(f'Processing Error: {ve}', 'danger')
             logging.error(f"ValueError during processing: {ve}")
        except FileNotFoundError as fnf:
             flash(f'File Error: {fnf}', 'danger')
             logging.error(f"FileNotFoundError during processing: {fnf}")
        except Exception as e:
            flash(f'An unexpected error occurred during processing. Please check file formats and try again.', 'danger')
            # Log the full error for debugging
            logging.error(f"Unexpected error during processing: {e}", exc_info=True) # Log traceback
        finally:
            # --- Clean up temporarily uploaded files ---
            logging.info(f"Cleaning up {len(saved_files)} uploaded files...")
            for sf in saved_files:
                if os.path.exists(sf):
                    try:
                        os.remove(sf)
                        logging.info(f"Cleaned up uploaded file: {sf}")
                    except OSError as e:
                        logging.warning(f"Could not remove uploaded file {sf} during cleanup: {e}")

        # Redirect back to form on any error encountered after file saving started
        return redirect(url_for('index'))

    # --- GET Request ---
    # Render the upload form
    return render_template('index.html')


# --- Main Execution ---
if __name__ == '__main__':
    # Development server (use Gunicorn for production/Render)
    # Host 0.0.0.0 makes it accessible on network (needed for Render health checks)
    # Use PORT environment variable provided by Render, default to 8080 locally
    port = int(os.environ.get("PORT", 8080))
    # Set debug=False for production environments like Render
    app.run(host='0.0.0.0', port=port, debug=False)
